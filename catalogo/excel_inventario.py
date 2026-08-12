import re
import unicodedata
from uuid import uuid4

from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import Autor, AutorRegistro, Editorial, Ejemplar, Materia, RegistroBibliografico

TEMPLATE_HEADERS = ['#', 'TITULO', 'AUTOR', 'EDITORIAL', 'EDICION', 'CANTIDAD', 'CATEGORIA', 'ISBN']

EJEMPLO_FILAS = [
    (1, 'Química General', 'Ebbing, Darrell D.', 'Cengage Learning', '10a edición', 3, 'Química', '9780538497527'),
    (2, 'Bioquímica', 'Lehninger, Albert L.', 'Omega', '6a edición', 2, 'Bioquímica', '9788428214704'),
    (3, 'Ejemplo Título', 'Apellido, Nombre', 'Editorial', '1a edición', 1, 'General', ''),
]


def generar_plantilla_xlsx() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Plantilla'
    ws.append(TEMPLATE_HEADERS)
    for fila in EJEMPLO_FILAS:
        ws.append(fila)
    return wb


def _normalizar(texto) -> str:
    texto = str(texto or '').strip().upper()
    texto = unicodedata.normalize('NFD', texto)
    return ''.join(c for c in texto if unicodedata.category(c) != 'Mn')


def importar_libros_xlsx(archivo) -> dict:
    """Importa libros desde un .xlsx. Ver catalogo/tests.py para casos cubiertos."""
    wb = load_workbook(archivo, read_only=True, data_only=True)
    ws = wb.active
    filas = [list(fila) for fila in ws.iter_rows(values_only=True)]

    fila_encabezado = None
    for i, fila in enumerate(filas[:5]):
        if any(re.match(r'^TITUL', _normalizar(c)) for c in fila if c is not None):
            fila_encabezado = i
            break
    if fila_encabezado is None:
        return {
            'nuevos': 0, 'actualizados': 0, 'errores': 0, 'resultados': [],
            'error': 'No se encontró la columna TITULO en el archivo.',
        }

    encabezados = [_normalizar(c) for c in filas[fila_encabezado]]

    def col(patron):
        for idx, h in enumerate(encabezados):
            if patron in h:
                return idx
        return -1

    i_titulo = col('TITUL')
    i_autor = col('AUTOR')
    i_editorial = encabezados.index('EDITORIAL') if 'EDITORIAL' in encabezados else -1
    i_edicion = next(
        (idx for idx, h in enumerate(encabezados) if 'EDIC' in h and 'ORIAL' not in h), -1
    )
    i_cantidad = col('CANTID')
    i_categoria = col('CATEG')
    i_isbn = col('ISBN')

    if i_titulo == -1 or i_autor == -1:
        return {
            'nuevos': 0, 'actualizados': 0, 'errores': 0, 'resultados': [],
            'error': 'El archivo debe tener columnas TITULO y AUTOR.',
        }

    def valor(fila, idx):
        if idx < 0 or idx >= len(fila) or fila[idx] is None:
            return ''
        return str(fila[idx]).strip()

    resultados = []
    nuevos = actualizados = errores = 0

    for n, fila in enumerate(filas[fila_encabezado + 1:], start=1):
        titulo = valor(fila, i_titulo)
        autor = valor(fila, i_autor)
        if not titulo or not autor:
            continue

        try:
            cantidad_raw = fila[i_cantidad] if 0 <= i_cantidad < len(fila) else None
            cantidad = max(1, int(cantidad_raw)) if cantidad_raw not in (None, '') else 1
        except (TypeError, ValueError):
            cantidad = 1
        editorial_nombre = valor(fila, i_editorial)
        edicion = valor(fila, i_edicion)
        categoria = valor(fila, i_categoria) or 'General'
        isbn = valor(fila, i_isbn)

        try:
            registro = RegistroBibliografico.objects.filter(
                titulo__iexact=titulo, autores__nombre__iexact=autor
            ).first()
            if registro:
                estado = 'actualizado'
                actualizados += 1
            else:
                editorial = None
                if editorial_nombre:
                    editorial, _ = Editorial.objects.get_or_create(nombre=editorial_nombre)
                registro = RegistroBibliografico.objects.create(
                    titulo=titulo, edicion=edicion, isbn=isbn, editorial=editorial,
                )
                autor_obj, _ = Autor.objects.get_or_create(nombre=autor)
                AutorRegistro.objects.create(registro=registro, autor=autor_obj, rol='PRINCIPAL', orden=1)
                materia, _ = Materia.objects.get_or_create(nombre=categoria)
                registro.materias.add(materia)
                estado = 'nuevo'
                nuevos += 1

            for _ in range(cantidad):
                Ejemplar.objects.create(registro=registro, codigo_barras=f'IMP-{uuid4().hex[:8].upper()}')

            resultados.append({'fila': n, 'estado': estado, 'titulo': titulo, 'autor': autor, 'cantidad': cantidad})
        except Exception as e:
            errores += 1
            resultados.append({'fila': n, 'estado': 'error', 'titulo': titulo, 'autor': autor, 'error': str(e)})

    return {'nuevos': nuevos, 'actualizados': actualizados, 'errores': errores, 'resultados': resultados}


def exportar_inventario_xlsx(tipo: str = 'todos') -> Workbook:
    registros = RegistroBibliografico.objects.select_related('editorial').prefetch_related(
        'autores', 'materias', 'ejemplares'
    )
    if tipo == 'este_mes':
        hoy = timezone.now()
        registros = registros.filter(fecha_alta__year=hoy.year, fecha_alta__month=hoy.month)
    elif tipo == 'ultima_carga':
        ultima = RegistroBibliografico.objects.order_by('-fecha_alta').values_list('fecha_alta', flat=True).first()
        registros = registros.filter(fecha_alta__date=ultima.date()) if ultima else registros.none()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    ws.append([
        'Título', 'Autor(es)', 'Editorial', 'Edición', 'Materias', 'ISBN',
        'Total ejemplares', 'Disponibles', 'Prestados', 'Fecha de alta',
    ])
    for registro in registros.order_by('titulo'):
        ejemplares = list(registro.ejemplares.all())
        disponibles = sum(1 for e in ejemplares if e.estado == 'DISPONIBLE')
        prestados = sum(1 for e in ejemplares if e.estado == 'PRESTADO')
        ws.append([
            registro.titulo,
            ', '.join(a.nombre for a in registro.autores.all()),
            registro.editorial.nombre if registro.editorial else '',
            registro.edicion,
            ', '.join(m.nombre for m in registro.materias.all()),
            registro.isbn,
            len(ejemplares), disponibles, prestados,
            registro.fecha_alta.strftime('%Y-%m-%d') if registro.fecha_alta else '',
        ])
    return wb
