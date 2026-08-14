import csv
import datetime
import io

import openpyxl
from django.contrib.auth.models import Group, Permission, User
from django.test import Client, TestCase
from django.urls import reverse
from django.utils import timezone
from pymarc.marcxml import parse_xml_to_array

from alumnos.models import Alumno
from circulacion.services import crear_prestamo

from .constancia_donacion import generar_constancia_donacion_docx
from .excel_inventario import (
    TEMPLATE_HEADERS, exportar_inventario_xlsx, generar_plantilla_xlsx, importar_libros_xlsx,
)
from .marc import _campo_008, generar_marc_xml
from .models import (
    Autor, AutorRegistro, ConstanciaDonacion, ConstanciaDonacionLibro,
    Editorial, Ejemplar, Materia, RegistroBibliografico,
)


def _xlsx_bytes(filas):
    wb = openpyxl.Workbook()
    ws = wb.active
    for fila in filas:
        ws.append(fila)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


class RegistroBibliograficoTests(TestCase):
    def setUp(self):
        self.autor = Autor.objects.create(nombre='Gabriel García Márquez')
        self.editorial = Editorial.objects.create(nombre='Sudamericana')
        self.materia = Materia.objects.create(nombre='Literatura latinoamericana')
        self.registro = RegistroBibliografico.objects.create(
            titulo='Cien años de soledad',
            editorial=self.editorial,
            clasificacion='863 G216c',
        )
        self.registro.materias.add(self.materia)
        AutorRegistro.objects.create(registro=self.registro, autor=self.autor, rol='PRINCIPAL')

    def test_registro_expone_autor_por_relacion_m2m(self):
        self.assertIn(self.autor, self.registro.autores.all())

    def test_registro_expone_materia(self):
        self.assertIn(self.materia, self.registro.materias.all())

    def test_no_se_puede_repetir_el_mismo_autor_dos_veces_en_un_registro(self):
        with self.assertRaises(Exception):
            AutorRegistro.objects.create(registro=self.registro, autor=self.autor, rol='COAUTOR')

    def test_ejemplar_queda_asociado_a_su_registro(self):
        ejemplar = Ejemplar.objects.create(registro=self.registro, codigo_barras='EJ-0001')
        self.assertIn(ejemplar, self.registro.ejemplares.all())
        self.assertEqual(ejemplar.estado, 'DISPONIBLE')


class MarcXmlTests(TestCase):
    def setUp(self):
        self.autor_principal = Autor.objects.create(nombre='García Márquez, Gabriel')
        self.coautor = Autor.objects.create(nombre='Vargas Llosa, Mario')
        self.editorial = Editorial.objects.create(nombre='Sudamericana')
        self.materia = Materia.objects.create(nombre='Literatura latinoamericana')
        self.registro = RegistroBibliografico.objects.create(
            titulo='Cien años de soledad',
            subtitulo='novela',
            isbn='978-0307474728',
            edicion='1a ed.',
            anio_publicacion=1967,
            lugar_publicacion='Buenos Aires, México',
            idioma='Español',
            descripcion_fisica='471 p. ; 21 cm.',
            clasificacion='863 G216c',
            notas='Incluye índice.',
            editorial=self.editorial,
        )
        self.registro.materias.add(self.materia)
        AutorRegistro.objects.create(registro=self.registro, autor=self.autor_principal, rol='PRINCIPAL', orden=1)
        AutorRegistro.objects.create(registro=self.registro, autor=self.coautor, rol='COAUTOR', orden=2)

    def test_campo_008_siempre_mide_40_caracteres(self):
        self.assertEqual(len(_campo_008(self.registro)), 40)

    def test_campo_008_mide_40_caracteres_con_datos_vacios(self):
        registro_vacio = RegistroBibliografico.objects.create(titulo='Sin datos')
        self.assertEqual(len(_campo_008(registro_vacio)), 40)

    def test_generar_marc_xml_de_registro_completo_es_legible_por_pymarc(self):
        xml = generar_marc_xml(self.registro)
        registros = parse_xml_to_array(io.BytesIO(xml.encode('utf-8')))

        self.assertEqual(len(registros), 1)
        record = registros[0]
        self.assertEqual(record['245']['a'], 'Cien años de soledad :')
        self.assertEqual(record['020']['a'], self.registro.isbn)

    def test_registro_con_solo_titulo_genera_xml_valido_sin_excepciones(self):
        registro_minimo = RegistroBibliografico.objects.create(titulo='Título mínimo')

        xml = generar_marc_xml(registro_minimo)
        registros = parse_xml_to_array(io.BytesIO(xml.encode('utf-8')))

        self.assertEqual(len(registros), 1)
        record = registros[0]
        self.assertEqual(record['245']['a'], 'Título mínimo')
        self.assertEqual(record.get_fields('020'), [])
        self.assertEqual(record.get_fields('082'), [])
        self.assertEqual(record.get_fields('250'), [])
        self.assertEqual(record.get_fields('300'), [])
        self.assertEqual(record.get_fields('500'), [])
        self.assertIn('[lugar no identificado]', record['264']['a'])
        self.assertIn('[editorial no identificada]', record['264']['b'])
        self.assertIn('[fecha no identificada]', record['264']['c'])

    def test_dos_autores_producen_exactamente_un_100_y_un_700_con_relator_correcto(self):
        xml = generar_marc_xml(self.registro)
        record = parse_xml_to_array(io.BytesIO(xml.encode('utf-8')))[0]

        campos_100 = record.get_fields('100')
        campos_700 = record.get_fields('700')

        self.assertEqual(len(campos_100), 1)
        self.assertEqual(campos_100[0]['a'], self.autor_principal.nombre)

        self.assertEqual(len(campos_700), 1)
        self.assertEqual(campos_700[0]['a'], self.coautor.nombre)
        self.assertEqual(campos_700[0]['e'], 'autor')


class BusquedaAdminTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_superuser('bibliotecario_test', 'biblio@example.com', 'x')
        self.client = Client()
        self.client.force_login(self.staff)

        autor = Autor.objects.create(nombre='Borges, Jorge Luis')
        editorial = Editorial.objects.create(nombre='Emecé')
        self.registro = RegistroBibliografico.objects.create(titulo='Ficciones', editorial=editorial)
        AutorRegistro.objects.create(registro=self.registro, autor=autor, rol='PRINCIPAL')

    def test_busqueda_por_nombre_de_autor_encuentra_el_registro(self):
        resp = self.client.get('/admin/catalogo/registrobibliografico/', {'q': 'Borges'})
        self.assertContains(resp, 'Ficciones')

    def test_busqueda_por_nombre_de_editorial_encuentra_el_registro(self):
        resp = self.client.get('/admin/catalogo/registrobibliografico/', {'q': 'Emecé'})
        self.assertContains(resp, 'Ficciones')

    def test_busqueda_sin_coincidencias_no_devuelve_el_registro(self):
        resp = self.client.get('/admin/catalogo/registrobibliografico/', {'q': 'inexistente_xyz'})
        self.assertNotContains(resp, 'Ficciones')


class EjemplarTests(TestCase):
    def setUp(self):
        self.registro = RegistroBibliografico.objects.create(titulo='Rayuela')

    def test_no_se_puede_repetir_el_mismo_codigo_de_barras_en_dos_ejemplares(self):
        Ejemplar.objects.create(registro=self.registro, codigo_barras='EJ-9000')
        with self.assertRaises(Exception):
            Ejemplar.objects.create(registro=self.registro, codigo_barras='EJ-9000')

    def test_ejemplar_se_puede_crear_desde_el_alta_inline_del_registro(self):
        staff = User.objects.create_superuser('bibliotecario_inline_test', 'biblio2@example.com', 'x')
        client = Client()
        client.force_login(staff)

        resp = client.get(f'/admin/catalogo/registrobibliografico/{self.registro.id}/change/')
        self.assertContains(resp, 'id_ejemplares-0-codigo_barras')


class PanelBibliotecaTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_superuser('bibliotecario_panel_test', 'biblio5@example.com', 'x')
        self.client = Client()
        self.client.force_login(self.staff)

        registro = RegistroBibliografico.objects.create(titulo='Pedro Páramo')
        self.ej_disponible = Ejemplar.objects.create(registro=registro, codigo_barras='EJ-0011')
        self.ej_prestado = Ejemplar.objects.create(registro=registro, codigo_barras='EJ-0012')
        alumno = Alumno.objects.create(
            nombre='ALUMNO PANEL TEST', numero_cuenta='9998893', carrera='MEDICO', facultad='MEDICINA'
        )
        ayer = timezone.now().date() - datetime.timedelta(days=1)
        crear_prestamo(
            alumno_nombre=alumno.nombre, matricula=alumno.numero_cuenta,
            fecha_devolucion=ayer, ejemplar_ids=[self.ej_prestado.pk],
        )

    def test_panel_biblioteca_muestra_las_estadisticas_esperadas(self):
        response = self.client.get(reverse('admin:catalogo_panel_biblioteca'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total_ejemplares'], 2)
        self.assertEqual(response.context['ejemplares_disponibles'], 1)
        self.assertEqual(response.context['prestamos_activos'], 1)
        self.assertEqual(response.context['prestamos_vencidos'], 1)


class EjemplarCsvExportTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_superuser('bibliotecario_csv_ej_test', 'biblio6@example.com', 'x')
        self.client = Client()
        self.client.force_login(self.staff)

        registro = RegistroBibliografico.objects.create(titulo='Ficciones')
        self.ejemplar = Ejemplar.objects.create(registro=registro, codigo_barras='EJ-0013', ubicacion='A-1')

    def test_exportar_csv_ejemplares_genera_csv_con_datos_correctos(self):
        response = self.client.post('/admin/catalogo/ejemplar/', {
            'action': 'exportar_ejemplares_csv',
            '_selected_action': [str(self.ejemplar.pk)],
        })
        self.assertEqual(response['Content-Type'], 'text/csv')
        self.assertIn('ejemplares.csv', response['Content-Disposition'])
        filas = list(csv.reader(io.StringIO(response.content.decode('utf-8'))))
        self.assertEqual(filas[0][0], 'Código de barras')
        self.assertEqual(filas[1][0], 'EJ-0013')
        self.assertEqual(filas[1][2], 'A-1')


def _texto_completo(doc):
    partes = [p.text for p in doc.paragraphs]
    for tabla in doc.tables:
        for fila in tabla.rows:
            for celda in fila.cells:
                partes.append(celda.text)
    return '\n'.join(partes)


class ConstanciaDonacionTests(TestCase):
    def setUp(self):
        self.autor = Autor.objects.create(nombre='Julio Cortázar')
        self.editorial = Editorial.objects.create(nombre='Sudamericana')
        self.registro = RegistroBibliografico.objects.create(
            titulo='Rayuela', editorial=self.editorial, edicion='1a ed.'
        )
        AutorRegistro.objects.create(registro=self.registro, autor=self.autor, rol='PRINCIPAL')

    def test_folio_tiene_padding_a_seis_digitos(self):
        constancia = ConstanciaDonacion.objects.create(persona_nombre='Juan Pérez')
        self.assertEqual(constancia.folio, f'CONST-{constancia.pk:06d}')

    def test_total_volumenes_suma_las_cantidades_de_cada_libro(self):
        constancia = ConstanciaDonacion.objects.create(persona_nombre='Juan Pérez')
        otro_registro = RegistroBibliografico.objects.create(titulo='Ficciones')
        ConstanciaDonacionLibro.objects.create(constancia=constancia, registro=self.registro, cantidad=2)
        ConstanciaDonacionLibro.objects.create(constancia=constancia, registro=otro_registro, cantidad=3)
        self.assertEqual(constancia.total_volumenes, 5)

    def test_generar_docx_produce_un_documento_valido_con_folio_donante_y_titulos(self):
        constancia = ConstanciaDonacion.objects.create(persona_nombre='María López', cargo='Directora')
        otro_registro = RegistroBibliografico.objects.create(titulo='Ficciones')
        ConstanciaDonacionLibro.objects.create(constancia=constancia, registro=self.registro, cantidad=1)
        ConstanciaDonacionLibro.objects.create(constancia=constancia, registro=otro_registro, cantidad=1)

        doc = generar_constancia_donacion_docx(constancia)
        texto = _texto_completo(doc)
        self.assertIn(constancia.folio, texto)
        self.assertIn('MARÍA LÓPEZ', texto)
        self.assertIn('Rayuela', texto)
        self.assertIn('Ficciones', texto)


class ConstanciaDonacionAdminAccessTests(TestCase):
    def setUp(self):
        self.registro = RegistroBibliografico.objects.create(titulo='Rayuela')
        self.constancia = ConstanciaDonacion.objects.create(persona_nombre='Juan Pérez')
        ConstanciaDonacionLibro.objects.create(constancia=self.constancia, registro=self.registro)

    def test_sin_permiso_view_constanciadonacion_no_puede_ver_el_changelist(self):
        usuario = User.objects.create_user('sin_permiso_constancia', 'x@example.com', 'x', is_staff=True)
        client = Client()
        client.force_login(usuario)
        response = client.get('/admin/catalogo/constanciadonacion/')
        self.assertEqual(response.status_code, 403)

    def test_con_permiso_view_constanciadonacion_puede_descargar_el_word(self):
        usuario = User.objects.create_user('con_permiso_constancia', 'y@example.com', 'x', is_staff=True)
        permisos = Permission.objects.filter(
            content_type__app_label='catalogo',
            codename__in=['view_constanciadonacion', 'change_constanciadonacion'],
        )
        usuario.user_permissions.add(*permisos)
        client = Client()
        client.force_login(usuario)

        response = client.post('/admin/catalogo/constanciadonacion/', {
            'action': 'descargar_word',
            '_selected_action': [str(self.constancia.pk)],
        })
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        )
        self.assertIn(f'{self.constancia.folio}.docx', response['Content-Disposition'])


class BibliotecarioPuedeCrearConstanciaConLibrosTests(TestCase):
    """El grupo Bibliotecario necesita permisos sobre ConstanciaDonacionLibro
    (no solo sobre ConstanciaDonacion) para que el inline del formulario de
    alta sea visible y guarde los libros seleccionados."""

    def setUp(self):
        self.usuario = User.objects.create_user('bibliotecario_grupo_test', 'z@example.com', 'x', is_staff=True)
        self.usuario.groups.add(Group.objects.get(name='Bibliotecario'))
        self.client = Client()
        self.client.force_login(self.usuario)
        self.registro = RegistroBibliografico.objects.create(titulo='El Aleph')

    def test_alta_de_constancia_con_libros_via_el_admin_guarda_los_renglones(self):
        response = self.client.post('/admin/catalogo/constanciadonacion/add/', {
            'persona_nombre': 'Donante de Prueba',
            'cargo': '',
            'fecha': timezone.now().date().isoformat(),
            'libros-TOTAL_FORMS': '1',
            'libros-INITIAL_FORMS': '0',
            'libros-MIN_NUM_FORMS': '0',
            'libros-MAX_NUM_FORMS': '1000',
            'libros-0-registro': str(self.registro.pk),
            'libros-0-cantidad': '2',
            '_save': 'Guardar',
        })
        self.assertEqual(response.status_code, 302)
        constancia = ConstanciaDonacion.objects.get(persona_nombre='Donante de Prueba')
        self.assertEqual(constancia.total_volumenes, 2)


class ImportarLibrosXlsxTests(TestCase):
    def test_libro_nuevo_crea_registro_autor_editorial_materia_y_ejemplares(self):
        archivo = _xlsx_bytes([
            ['#', 'TITULO', 'AUTOR', 'EDITORIAL', 'EDICION', 'CANTIDAD', 'CATEGORIA', 'ISBN'],
            [1, 'Nuevo Libro', 'Autor Nuevo', 'Editorial X', '1a ed.', 3, 'Categoria X', '123456'],
        ])
        resultado = importar_libros_xlsx(archivo)
        self.assertEqual(resultado['nuevos'], 1)
        self.assertEqual(resultado['actualizados'], 0)
        self.assertEqual(resultado['errores'], 0)

        registro = RegistroBibliografico.objects.get(titulo='Nuevo Libro')
        self.assertEqual(registro.ejemplares.count(), 3)
        self.assertEqual(registro.editorial.nombre, 'Editorial X')
        self.assertIn('Autor Nuevo', [a.nombre for a in registro.autores.all()])
        self.assertIn('Categoria X', [m.nombre for m in registro.materias.all()])
        self.assertTrue(all(e.codigo_barras.startswith('IMP-') for e in registro.ejemplares.all()))

    def test_libro_existente_agrega_ejemplares_sin_duplicar_registro(self):
        registro = RegistroBibliografico.objects.create(titulo='Libro Existente')
        autor = Autor.objects.create(nombre='Autor Existente')
        AutorRegistro.objects.create(registro=registro, autor=autor, rol='PRINCIPAL')
        Ejemplar.objects.create(registro=registro, codigo_barras='EJ-EXIST-1')

        archivo = _xlsx_bytes([
            ['TITULO', 'AUTOR', 'CANTIDAD'],
            ['Libro Existente', 'Autor Existente', 2],
        ])
        resultado = importar_libros_xlsx(archivo)
        self.assertEqual(resultado['nuevos'], 0)
        self.assertEqual(resultado['actualizados'], 1)
        self.assertEqual(RegistroBibliografico.objects.filter(titulo='Libro Existente').count(), 1)
        registro.refresh_from_db()
        self.assertEqual(registro.ejemplares.count(), 3)

    def test_encabezados_en_otro_orden_se_detectan_por_nombre(self):
        archivo = _xlsx_bytes([
            ['AUTOR', 'CANTIDAD', 'TITULO'],
            ['Autor Orden', 1, 'Libro Orden'],
        ])
        resultado = importar_libros_xlsx(archivo)
        self.assertEqual(resultado['nuevos'], 1)
        self.assertTrue(RegistroBibliografico.objects.filter(titulo='Libro Orden').exists())

    def test_fila_sin_titulo_o_autor_se_ignora(self):
        archivo = _xlsx_bytes([
            ['TITULO', 'AUTOR'],
            ['', 'Autor sin título'],
            ['Libro sin autor', ''],
            ['Libro Válido', 'Autor Válido'],
        ])
        resultado = importar_libros_xlsx(archivo)
        self.assertEqual(resultado['nuevos'], 1)
        self.assertEqual(len(resultado['resultados']), 1)

    def test_archivo_sin_columna_titulo_devuelve_error_controlado(self):
        archivo = _xlsx_bytes([
            ['AUTOR', 'CANTIDAD'],
            ['Autor X', 1],
        ])
        resultado = importar_libros_xlsx(archivo)
        self.assertIn('error', resultado)
        self.assertEqual(resultado['nuevos'], 0)


class ExcelInventarioGeneracionTests(TestCase):
    def test_generar_plantilla_produce_workbook_con_encabezados_esperados(self):
        wb = generar_plantilla_xlsx()
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        wb_leido = openpyxl.load_workbook(buffer)
        encabezados = [c.value for c in next(wb_leido.active.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(encabezados, TEMPLATE_HEADERS)

    def test_exportar_inventario_incluye_registros_existentes(self):
        registro = RegistroBibliografico.objects.create(titulo='Libro Exportado')
        Ejemplar.objects.create(registro=registro, codigo_barras='EJ-EXPORT-1')

        wb = exportar_inventario_xlsx('todos')
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        wb_leido = openpyxl.load_workbook(buffer)
        titulos = [fila[0].value for fila in wb_leido.active.iter_rows(min_row=2)]
        self.assertIn('Libro Exportado', titulos)


class PanelBibliotecaExcelVistasAccesoTests(TestCase):
    def setUp(self):
        self.registro = RegistroBibliografico.objects.create(titulo='Libro Acceso')
        Ejemplar.objects.create(registro=self.registro, codigo_barras='EJ-ACCESO-1')

    def _cliente_sin_permiso(self, username):
        usuario = User.objects.create_user(username, f'{username}@example.com', 'x', is_staff=True)
        client = Client()
        client.force_login(usuario)
        return client

    def _cliente_con_permiso(self, username):
        usuario = User.objects.create_user(username, f'{username}@example.com', 'x', is_staff=True)
        permiso = Permission.objects.get(
            content_type__app_label='catalogo', codename='view_registrobibliografico'
        )
        usuario.user_permissions.add(permiso)
        client = Client()
        client.force_login(usuario)
        return client

    def test_sin_permiso_no_puede_acceder_a_importar(self):
        client = self._cliente_sin_permiso('sin_permiso_importar')
        response = client.get(reverse('admin:catalogo_panel_biblioteca_importar'))
        self.assertEqual(response.status_code, 403)

    def test_sin_permiso_no_puede_acceder_a_plantilla(self):
        client = self._cliente_sin_permiso('sin_permiso_plantilla')
        response = client.get(reverse('admin:catalogo_panel_biblioteca_plantilla'))
        self.assertEqual(response.status_code, 403)

    def test_sin_permiso_no_puede_acceder_a_exportar(self):
        client = self._cliente_sin_permiso('sin_permiso_exportar')
        response = client.get(reverse('admin:catalogo_panel_biblioteca_exportar'))
        self.assertEqual(response.status_code, 403)

    def test_sin_permiso_no_puede_acceder_a_imprimir(self):
        client = self._cliente_sin_permiso('sin_permiso_imprimir')
        response = client.get(reverse('admin:catalogo_panel_biblioteca_imprimir'))
        self.assertEqual(response.status_code, 403)

    def test_con_permiso_puede_descargar_plantilla(self):
        client = self._cliente_con_permiso('con_permiso_plantilla')
        response = client.get(reverse('admin:catalogo_panel_biblioteca_plantilla'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_con_permiso_puede_exportar_e_imprimir(self):
        client = self._cliente_con_permiso('con_permiso_exportar')
        response = client.get(reverse('admin:catalogo_panel_biblioteca_exportar'), {'tipo': 'todos'})
        self.assertEqual(response.status_code, 200)

        response = client.get(reverse('admin:catalogo_panel_biblioteca_imprimir'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Libro Acceso')

    def test_con_permiso_puede_ver_formulario_de_importar(self):
        client = self._cliente_con_permiso('con_permiso_importar')
        response = client.get(reverse('admin:catalogo_panel_biblioteca_importar'))
        self.assertEqual(response.status_code, 200)
